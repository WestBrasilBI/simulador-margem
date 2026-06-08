"""
Roda diariamente para atualizar os dados do app a partir da planilha.
Uso: python convert_excel.py
"""
import openpyxl
import json
import os
import sys

FILE_PATH = r"C:\Users\usuario\OneDrive - WEST BRASIL LUBRIFICANTES\RepositorioBI - ArquivosBI\Base Excel\Rafa\SIMULADOR MARGEM UNITÁRIA - OFICIAL.xlsx"
OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "data.js")

print("Abrindo planilha...")
try:
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True, read_only=True)
except FileNotFoundError:
    print(f"ERRO: Arquivo não encontrado em {FILE_PATH}")
    sys.exit(1)

# ── CLIENTES ─────────────────────────────────────────────────────────────────
print("Extraindo clientes...")
ws = wb['BASE CLIENTES']
clientes = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    cod = row[2]
    if not cod:
        continue
    try:
        key = str(int(cod))
    except (ValueError, TypeError):
        key = str(cod).strip()
    imbp_sim = row[13] if row[13] else (row[10] if row[10] else "")
    clientes[key] = {
        "nome":      str(row[5] or "").strip(),
        "segmento":  str(row[6] or "").strip(),
        "ramo":      str(row[9] or "").strip(),
        "imbp":      str(imbp_sim).strip(),
    }
print(f"  {len(clientes)} clientes")

# ── CUSTO ────────────────────────────────────────────────────────────────────
print("Extraindo tabela de custos...")
ws = wb['custo']
custo = {}
nomes_produto = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    cod  = row[0]
    nome = row[1]
    preco = row[6]
    cst  = row[7]
    concat = row[9]
    if concat is None or cod is None:
        continue
    key = str(concat).strip()
    custo[key] = {
        "cod":        str(cod).strip(),
        "nome":       str(nome or "").strip(),
        "custo":      float(cst) if cst else 0.0,
        "preco_lista": float(preco) if preco else 0.0,
    }
    # mapa simples cod -> nome para autocomplete
    cod_str = str(cod).strip()
    if cod_str not in nomes_produto:
        nomes_produto[cod_str] = str(nome or "").strip()
print(f"  {len(custo)} entradas custo, {len(nomes_produto)} SKUs únicos")

# ── CONVERSÃO ────────────────────────────────────────────────────────────────
print("Extraindo fatores de conversão...")
ws = wb['conversão']
conversao = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    cod   = row[0]
    ativo = row[3]
    fator = row[6]
    if cod is None or not fator:
        continue
    cod_str = str(cod).strip()
    # prefere entrada ativa; só sobrescreve se ainda não tem entrada ativa
    if cod_str not in conversao or ativo == 1:
        conversao[cod_str] = float(fator)
print(f"  {len(conversao)} produtos com fator de conversão")

# ── PREMIER ──────────────────────────────────────────────────────────────────
print("Extraindo tabela Premier...")
ws = wb['PREMIER']
premier_default = 0.609
# Linha 1 contém CMV NF
for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
    if row and len(row) > 7 and row[6] == 'CMV NF':
        try:
            premier_default = float(row[7])
        except (TypeError, ValueError):
            pass

premier = {}
for row in ws.iter_rows(min_row=5, values_only=True):
    concat_total = row[5]  # coluna F – CONCAT TOTAL
    pct = row[7]           # coluna H – %
    if concat_total and pct is not None:
        try:
            premier[str(concat_total).strip()] = float(pct)
        except (ValueError, TypeError):
            pass
print(f"  {len(premier)} entradas Premier, default CMV NF = {premier_default:.4f}")

# ── AUX (prazos) ─────────────────────────────────────────────────────────────
print("Extraindo prazos de pagamento...")
ws = wb['AUX']
aux = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    prazo = row[0]
    pct   = row[1]
    if prazo and pct is not None:
        aux[str(prazo).strip()] = float(pct)
print(f"  {len(aux)} prazos: {list(aux.keys())}")

wb.close()

# ── GRAVA data.js ─────────────────────────────────────────────────────────────
print(f"Gravando {OUTPUT_PATH}...")
with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    f.write("// Gerado automaticamente por convert_excel.py — não editar manualmente\n")
    f.write(f'const DATA_VERSION = "{__import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M")}";\n')
    f.write("const CLIENTES = "); json.dump(clientes, f, ensure_ascii=False); f.write(";\n")
    f.write("const CUSTO = ");    json.dump(custo, f, ensure_ascii=False);    f.write(";\n")
    f.write("const CONVERSAO = "); json.dump(conversao, f, ensure_ascii=False); f.write(";\n")
    f.write("const PREMIER = ");  json.dump(premier, f, ensure_ascii=False);  f.write(";\n")
    f.write(f"const PREMIER_DEFAULT = {premier_default};\n")
    f.write("const AUX = ");      json.dump(aux, f, ensure_ascii=False);      f.write(";\n")
    f.write("const NOMES_PRODUTO = "); json.dump(nomes_produto, f, ensure_ascii=False); f.write(";\n")

size_kb = os.path.getsize(OUTPUT_PATH) / 1024
print(f"Concluído! data.js gerado ({size_kb:.0f} KB)")
